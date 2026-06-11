# This code is built by Engr. Brian Ezekiel D. Batalon, ECT, SO2
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, redirect, url_for
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sendgrid
from sendgrid.helpers.mail import Mail, Email, To, Content
import json
import os
from dotenv import load_dotenv
load_dotenv()
import re

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'grademaster-secret-key-2026')
db = SQLAlchemy(app)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# SendGrid Configuration
SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'aztechworx@gmail.com')

# Database Models
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    display_name = db.Column(db.String(200), default='')
    sections = db.relationship('Section', backref='owner', lazy=True, cascade='all, delete-orphan')

class Section(db.Model):
    __tablename__ = 'sections'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    passing_grade = db.Column(db.Integer, default=75)
    rubrics = db.Column(db.Text, default='[]')
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    students = db.relationship('Student', backref='section', lazy=True, cascade='all, delete-orphan')
    attendance_records = db.relationship('Attendance', backref='section', lazy=True, cascade='all, delete-orphan')

class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(200), default='')
    scores = db.Column(db.Text, default='{}')
    final_grade = db.Column(db.Float, nullable=True)
    grade_point = db.Column(db.Float, nullable=True)
    remark = db.Column(db.String(50), default='')
    status = db.Column(db.String(20), default='')
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)

class Attendance(db.Model):
    __tablename__ = 'attendance'
    id = db.Column(db.Integer, primary_key=True)
    section_id = db.Column(db.Integer, db.ForeignKey('sections.id'), nullable=False)
    date = db.Column(db.String(50), nullable=False)
    records = db.Column(db.Text, default='{}')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Score Parsing
def parse_score(value, total_items=None):
    if value is None or value == '':
        return ''
    if isinstance(value, str) and value.strip().upper() in ['INC', 'AD', 'UD']:
        return value.strip().upper()
    if isinstance(value, str) and '/' in value:
        parts = value.split('/')
        try:
            score = float(parts[0].strip())
            total = float(parts[1].strip())
            if total == 0:
                return ''
            return round((score / total) * 100, 2)
        except (ValueError, ZeroDivisionError):
            return value
    try:
        num = float(value)
        if total_items is not None and total_items > 0 and num <= total_items:
            return round((num / total_items) * 100, 2)
        return num
    except (ValueError, TypeError):
        return value

# Grade Translation
def translate_grade(score):
    mapping = [
        (100, 1.00, "Excellent"),
        (99, 1.10, "Excellent"),
        (98, 1.10, "Excellent"),
        (97, 1.20, "Excellent"),
        (96, 1.20, "Excellent"),
        (95, 1.30, "Very Good"),
        (94, 1.30, "Very Good"),
        (93, 1.40, "Very Good"),
        (92, 1.40, "Very Good"),
        (91, 1.50, "Very Good"),
        (90, 1.50, "Very Good"),
        (89, 1.60, "Very Good"),
        (88, 1.70, "Very Good"),
        (87, 1.70, "Very Good"),
        (86, 1.80, "Good"),
        (85, 2.00, "Good"),
        (84, 2.10, "Good"),
        (83, 2.20, "Fair"),
        (82, 2.30, "Fair"),
        (81, 2.40, "Fair"),
        (80, 2.50, "Fair"),
        (79, 2.60, "Fair"),
        (78, 2.70, "Fair"),
        (77, 2.80, "Fair"),
        (76, 2.90, "Fair"),
        (75, 3.00, "Passing"),
    ]
    if score is None or score == '':
        return {"grade_point": None, "remark": None}
    for threshold, gp, remark in mapping:
        if score >= threshold:
            return {"grade_point": gp, "remark": remark}
    return {"grade_point": 5.00, "remark": "Failed"}

# ─── AUTH ROUTES ───
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return jsonify({'status': 'success', 'user': {'username': user.username, 'display_name': user.display_name}})
        return jsonify({'status': 'error', 'message': 'Invalid username or password'}), 401
    return render_template('login.html')

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    display_name = data.get('display_name', username)
    if User.query.filter_by(username=username).first():
        return jsonify({'status': 'error', 'message': 'Username already exists'}), 400
    if len(password) < 4:
        return jsonify({'status': 'error', 'message': 'Password must be at least 4 characters'}), 400
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        display_name=display_name
    )
    db.session.add(user)
    db.session.commit()
    login_user(user)
    return jsonify({'status': 'success', 'user': {'username': user.username, 'display_name': user.display_name}})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/current_user')
@login_required
def get_current_user():
    return jsonify({'username': current_user.username, 'display_name': current_user.display_name})

# ─── PROTECTED ROUTES ───
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.ico', mimetype='image/x-icon')

@app.route('/api/sections', methods=['GET', 'POST', 'DELETE'])
@login_required
def manage_sections():
    if request.method == 'POST':
        data = request.json
        name = data['name'].strip()
        if Section.query.filter_by(name=name).first():
            return jsonify({'status': 'error', 'message': 'Section already exists'}), 400
        section = Section(name=name, user_id=current_user.id)
        db.session.add(section)
        db.session.commit()
        return jsonify({'status': 'success', 'sections': [s.name for s in Section.query.filter_by(user_id=current_user.id).all()]})
    elif request.method == 'DELETE':
        name = request.json.get('name')
        section = Section.query.filter_by(name=name, user_id=current_user.id).first()
        if section:
            db.session.delete(section)
            db.session.commit()
        return jsonify({'status': 'success'})
    return jsonify([s.name for s in Section.query.filter_by(user_id=current_user.id).all()])

@app.route('/api/sections/<section_name>/rubrics', methods=['GET', 'POST'])
@login_required
def manage_rubrics(section_name):
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section:
        return jsonify({'status': 'error', 'message': 'Section not found'}), 404
    if request.method == 'POST':
        data = request.json
        section.rubrics = json.dumps(data.get('rubrics', []))
        section.passing_grade = data.get('passing_grade', 75)
        db.session.commit()
        return jsonify({'status': 'success', 'section': section_to_dict(section)})
    return jsonify(section_to_dict(section))

@app.route('/api/sections/<section_name>/students', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def manage_students(section_name):
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section:
        return jsonify({'status': 'error', 'message': 'Section not found'}), 404
    if request.method == 'POST':
        student_data = request.json
        student = Student(
            first_name=student_data.get('first_name', '').strip(),
            last_name=student_data.get('last_name', '').strip(),
            email=student_data.get('email', '').strip(),
            scores=json.dumps(student_data.get('scores', {})),
            section_id=section.id
        )
        db.session.add(student)
        db.session.commit()
        compute_student_grades(student, section)
        db.session.commit()
        return jsonify({'status': 'success', 'student': student_to_dict(student)})
    elif request.method == 'PUT':
        data = request.json
        index = data.pop('index', None)
        students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
        if index is not None and 0 <= index < len(students_list):
            student = students_list[index]
            student.first_name = data.get('first_name', student.first_name)
            student.last_name = data.get('last_name', student.last_name)
            student.email = data.get('email', student.email)
            student.scores = json.dumps(data.get('scores', {}))
            db.session.commit()
            compute_student_grades(student, section)
            db.session.commit()
            return jsonify({'status': 'success', 'student': student_to_dict(student)})
        return jsonify({'status': 'error'}), 400
    elif request.method == 'DELETE':
        index = request.json.get('index')
        students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
        if index is not None and 0 <= index < len(students_list):
            db.session.delete(students_list[index])
            db.session.commit()
        return jsonify({'status': 'success'})
    students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
    return jsonify([student_to_dict(s) for s in students_list])

def compute_student_grades(student, section):
    rubrics = json.loads(section.rubrics) if section.rubrics else []
    scores = json.loads(student.scores) if student.scores else {}
    total = 0
    for rubric in rubrics:
        r_name = rubric['name']
        weight = rubric['weight']
        if 'items' in rubric and rubric['items']:
            sub_total = 0
            count = 0
            all_null = True
            for item in rubric['items']:
                item_name = item['name'] if isinstance(item, dict) else item
                item_total = item.get('total_items', None) if isinstance(item, dict) else None
                raw_score = scores.get(f"{r_name}::{item_name}")
                score = parse_score(raw_score, item_total)
                if score != '' and score not in ['INC', 'AD', 'UD'] and isinstance(score, (int, float)):
                    sub_total += float(score)
                    count += 1
                    all_null = False
                elif score in ['INC', 'AD', 'UD']:
                    all_null = False
            if not all_null and count > 0:
                avg = sub_total / count
                total += (avg * weight) / 100
        else:
            raw_score = scores.get(r_name)
            score = parse_score(raw_score)
            if score != '' and score not in ['INC', 'AD', 'UD'] and isinstance(score, (int, float)):
                total += (float(score) * weight) / 100
    has_inc = any(parse_score(v) == 'INC' for v in scores.values())
    has_ad = any(parse_score(v) == 'AD' for v in scores.values())
    has_ud = any(parse_score(v) == 'UD' for v in scores.values())
    if has_ud:
        student.status = 'UD'; student.grade_point = None; student.remark = 'Unauthorized Dropped'; student.final_grade = None
    elif has_ad:
        student.status = 'AD'; student.grade_point = None; student.remark = 'Authorized Dropped'; student.final_grade = None
    elif has_inc:
        student.status = 'INC'; student.grade_point = None; student.remark = 'Incomplete'; student.final_grade = None
    else:
        student.final_grade = round(total, 2)
        translation = translate_grade(round(total))
        student.grade_point = translation['grade_point']
        student.remark = translation['remark']
        student.status = 'PASSED' if total >= section.passing_grade else 'FAILED'

@app.route('/api/sections/<section_name>/export')
@login_required
def export_section(section_name):
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section:
        return jsonify({'status': 'error'}), 404
    rubrics = json.loads(section.rubrics) if section.rubrics else []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = section_name
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    headers = ['Last Name', 'First Name', 'Email']
    for rubric in rubrics:
        if 'items' in rubric and rubric['items']:
            for item in rubric['items']:
                item_name = item['name'] if isinstance(item, dict) else item
                headers.append(f"{rubric['name']} - {item_name}")
        else:
            headers.append(f"{rubric['name']} ({rubric['weight']}%)")
    headers += ['Final Grade', 'Grade Point', 'Remark', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
    for row, student in enumerate(students_list, 2):
        scores = json.loads(student.scores) if student.scores else {}
        ws.cell(row=row, column=1, value=student.last_name).border = thin_border
        ws.cell(row=row, column=2, value=student.first_name).border = thin_border
        ws.cell(row=row, column=3, value=student.email or '').border = thin_border
        col = 4
        for rubric in rubrics:
            if 'items' in rubric and rubric['items']:
                for item in rubric['items']:
                    item_name = item['name'] if isinstance(item, dict) else item
                    score = scores.get(f"{rubric['name']}::{item_name}", '')
                    ws.cell(row=row, column=col, value=score).border = thin_border
                    ws.cell(row=row, column=col).alignment = center_align
                    col += 1
            else:
                score = scores.get(rubric['name'], '')
                ws.cell(row=row, column=col, value=score).border = thin_border
                ws.cell(row=row, column=col).alignment = center_align
                col += 1
        ws.cell(row=row, column=col, value=student.final_grade or '').border = thin_border
        ws.cell(row=row, column=col).alignment = center_align
        ws.cell(row=row, column=col+1, value=student.grade_point or '').border = thin_border
        ws.cell(row=row, column=col+1).alignment = center_align
        ws.cell(row=row, column=col+2, value=student.remark or '').border = thin_border
        ws.cell(row=row, column=col+2).alignment = center_align
        status_cell = ws.cell(row=row, column=col+3, value=student.status or '')
        status_cell.border = thin_border; status_cell.alignment = center_align
        if student.status == 'PASSED': status_cell.font = Font(color="008000", bold=True)
        elif student.status == 'FAILED': status_cell.font = Font(color="FF0000", bold=True)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value: max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 6, 30)
    output = BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{section_name}_grades.xlsx')

@app.route('/api/sections/<section_name>/email', methods=['POST'])
@login_required
def email_student(section_name):
    data = request.json
    index = data.get('index')
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section: return jsonify({'status': 'error'}), 404
    students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
    if index is None or index >= len(students_list): return jsonify({'status': 'error'}), 404
    student = students_list[index]
    if not student.email: return jsonify({'status': 'error', 'message': 'No email'}), 400
    subject = f"Your Grade for {section_name}"
    body = f"""<h2>Grade Report</h2><p><strong>Student:</strong> {student.last_name}, {student.first_name}</p><p><strong>Section:</strong> {section_name}</p><hr>
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse:collapse;">
    <tr><td><strong>Final Grade:</strong></td><td>{student.final_grade if student.final_grade is not None else 'N/A'}%</td></tr>
    <tr><td><strong>Grade Point:</strong></td><td>{student.grade_point if student.grade_point is not None else 'N/A'}</td></tr>
    <tr><td><strong>Remark:</strong></td><td>{student.remark or 'N/A'}</td></tr>
    <tr><td><strong>Status:</strong></td><td>{student.status or 'N/A'}</td></tr></table>"""
    try:
        sg = sendgrid.SendGridAPIClient(api_key=SENDGRID_API_KEY)
        message = Mail(
            from_email=Email(SENDER_EMAIL, "GradeMaster"),
            to_emails=To(student.email),
            subject=subject,
            html_content=body
        )
        response = sg.send(message)
        return jsonify({'status': 'success', 'message': f'Email sent to {student.email}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ─── ATTENDANCE ───
@app.route('/api/sections/<section_name>/attendance', methods=['GET', 'POST'])
@login_required
def manage_attendance(section_name):
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section: return jsonify({'status': 'error'}), 404
    if request.method == 'POST':
        data = request.json; date = data.get('date', '').strip(); records = data.get('records', {})
        existing = Attendance.query.filter_by(section_id=section.id, date=date).first()
        if existing: existing.records = json.dumps(records)
        else: db.session.add(Attendance(section_id=section.id, date=date, records=json.dumps(records)))
        db.session.commit()
        return jsonify({'status': 'success'})
    records = Attendance.query.filter_by(section_id=section.id).order_by(Attendance.date).all()
    return jsonify([{'date': r.date, 'records': json.loads(r.records) if r.records else {}} for r in records])

@app.route('/api/sections/<section_name>/attendance/<date>', methods=['DELETE'])
@login_required
def delete_attendance(section_name, date):
    section = Section.query.filter_by(name=section_name, user_id=current_user.id).first()
    if not section: return jsonify({'status': 'error'}), 404
    record = Attendance.query.filter_by(section_id=section.id, date=date).first()
    if record: db.session.delete(record); db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/api/reset')
@login_required
def reset():
    Attendance.query.filter(Attendance.section.has(user_id=current_user.id)).delete(synchronize_session=False)
    Student.query.filter(Student.section.has(user_id=current_user.id)).delete(synchronize_session=False)
    Section.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    return jsonify({'status': 'success'})

# Helper Functions
def section_to_dict(section):
    students_list = sorted(section.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))
    return {'name': section.name, 'passing_grade': section.passing_grade,
            'rubrics': json.loads(section.rubrics) if section.rubrics else [],
            'students': [student_to_dict(s) for s in students_list]}

def student_to_dict(student):
    return {'first_name': student.first_name, 'last_name': student.last_name, 'email': student.email or '',
            'scores': json.loads(student.scores) if student.scores else {},
            'final_grade': student.final_grade, 'grade_point': student.grade_point,
            'remark': student.remark or '', 'status': student.status or ''}

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)